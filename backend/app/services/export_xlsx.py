"""Excel export service — five workbook flavors for the GRX10 market research platform.

Every workbook produced here:

* Opens with a ``_README`` sheet at position 0: filter scope, export timestamp,
  methodology hyperlink, and the four spec invariants so any recipient can trace
  the methodology without reading the codebase.
* Hyperlinks source ``url_pattern`` values in all "Source URL" columns (blue,
  underlined, HYPERLINK formula compatible).
* Confidence cells are colour-coded: HIGH=green, MEDIUM=amber, LOW=red.
* Column widths are auto-sized from content (capped at 60 chars).

Public surface
--------------
:func:`build_workbook`
    Dispatch function — returns the correct openpyxl ``Workbook`` for *flavor*.

:func:`workbook_to_bytes`
    Save a ``Workbook`` to ``BytesIO`` and return the raw bytes for streaming.

Flavors
-------
``cell_explorer``
    Filterable cells (subcategory × geography × year), TAM band (low / revenue /
    high in $M), confidence chip.  One data sheet: *Cells*.

``cell_detail``
    Full audit drill chain: cell summary + one row per triangulation estimate,
    embedding method (tier, cap), source (publisher, URL), raw-table pointer.
    Two data sheets: *Cell Summary*, *Estimates*.

``player_shares``
    Ranked company market shares per cell + buyer-supplier relationship edges.
    Two data sheets: *Player Shares*, *Supplier Relationships*.

``triangulation``
    Per-method estimates + confidence-math projection from the
    ``cell_triangulation_summary`` materialised view.
    Two data sheets: *Triangulation*, *Tri Summary*.

``assumptions``
    Versioned assumption ledger (with ``superseded_by`` chain) + the
    ``cell_assumption_link`` bridge table showing influenced cells.
    Two data sheets: *Assumptions*, *Cell Links*.
"""

from __future__ import annotations

import io
import logging
from datetime import datetime, timezone
from decimal import Decimal
from typing import Any, Literal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session, joinedload

from backend.app.config import settings
from backend.app.models import (
    Assumption,
    Cell,
    CellAssumptionLink,
    CellTriangulation,
    CellTriangulationSummary,
    Company,
    PlayerShare,
    Source,
    SupplierRelationship,
)

logger = logging.getLogger("grx10.services.export_xlsx")

ExportFlavor = Literal[
    "cell_explorer", "cell_detail", "player_shares", "triangulation", "assumptions"
]

# =========================================================================== #
# Style constants
# =========================================================================== #
_HEADER_FONT   = Font(bold=True, color="FFFFFF", size=10)
_HEADER_FILL   = PatternFill("solid", fgColor="2D6A9F")  # GRX10 blue
_HYPER_FONT    = Font(color="0563C1", underline="single", size=10)
_LABEL_FONT    = Font(bold=True, size=10)
_SECTION_FONT  = Font(bold=True, color="2D6A9F", size=10)
_TITLE_FONT    = Font(bold=True, size=13)
_BODY_ALIGN    = Alignment(wrap_text=True, vertical="top")

_CONFIDENCE_FILL: dict[str, PatternFill] = {
    "high":   PatternFill("solid", fgColor="C6EFCE"),  # green
    "medium": PatternFill("solid", fgColor="FFEB9C"),  # amber
    "low":    PatternFill("solid", fgColor="FFC7CE"),  # red
}

_COL_MIN = 8
_COL_MAX = 60


# =========================================================================== #
# Low-level worksheet helpers
# =========================================================================== #

def _write_headers(ws, headers: list[str]) -> dict[int, int]:
    """Write bold/coloured header row, freeze row 1, and return a width-tracker."""
    widths: dict[int, int] = {}
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        widths[col]    = len(label)
    ws.freeze_panes = "A2"
    return widths


def _set_hyperlink(cell, url: str | None, display: str | None = None) -> None:
    """Make *cell* a clickable hyperlink; no-op when *url* is falsy."""
    if not url:
        cell.value = display or ""
        return
    cell.value     = display or url
    cell.hyperlink = url
    cell.font      = _HYPER_FONT


def _track(widths: dict[int, int], col: int, value: Any) -> None:
    """Update the per-column content-length tracker."""
    length = len(str(value)) if value is not None else 0
    if length > widths.get(col, 0):
        widths[col] = length


def _apply_widths(ws, widths: dict[int, int]) -> None:
    """Set column widths from the tracker (clamped between min and max)."""
    for col, length in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(
            max(length + 2, _COL_MIN), _COL_MAX
        )


def _write_row(
    ws,
    row_idx: int,
    values: list[Any],
    widths: dict[int, int],
    *,
    hyperlink_cols: dict[int, str | None] | None = None,
    confidence_col: int | None = None,
) -> None:
    """Write one data row; handle hyperlinks and confidence colouring.

    Parameters
    ----------
    values:
        Ordered cell values for columns 1..N.
    hyperlink_cols:
        ``{col_idx: display_text}`` where the cell value carries the raw URL.
        Pass ``None`` as display_text to fall back to the URL itself.
    confidence_col:
        Column index of the confidence field; that cell gets a conditional fill.
    """
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col)

        # Strip timezone from datetimes — openpyxl does not support tz-aware
        # datetimes and raises during wb.save(); strip here before assignment.
        if isinstance(value, datetime) and value.tzinfo is not None:
            value = value.replace(tzinfo=None)

        if hyperlink_cols and col in hyperlink_cols:
            display = hyperlink_cols[col]
            _set_hyperlink(cell, str(value) if value else None, display)
        else:
            cell.value = value

        _track(widths, col, value)

        if confidence_col and col == confidence_col:
            fill = _CONFIDENCE_FILL.get(str(value) if value else "")
            if fill:
                cell.fill = fill


# =========================================================================== #
# README sheet
# =========================================================================== #

def _add_readme_sheet(
    wb: Workbook,
    flavor: str,
    filter_desc: dict[str, Any],
    ts: datetime,
    methodology_url: str,
) -> None:
    """Insert a ``_README`` sheet at workbook position 0.

    Records: export flavor, UTC timestamp, applied filters, methodology hyperlink,
    and the four spec invariants so any downstream consumer can trace the data
    provenance without reading source code.
    """
    ws = wb.create_sheet("_README", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 68

    # Build a flat list of (row_label, value, style_hint) tuples.
    # style_hint: "title" | "section" | "field" | None (blank row)
    entries: list[tuple[str | None, Any, str | None, bool]] = []
    # (label, value, style, is_hyperlink)

    entries.append(("GRX10 Market Research Tool — Export", None, "title", False))
    entries.append((None, None, None, False))
    entries.append(("Export flavor", flavor.replace("_", " ").title(), "field", False))
    entries.append(("Export timestamp (UTC)", ts.strftime("%Y-%m-%dT%H:%M:%SZ"), "field", False))
    entries.append((None, None, None, False))

    entries.append(("FILTER SCOPE", None, "section", False))
    active_filters = {k: v for k, v in (filter_desc or {}).items() if v is not None}
    if active_filters:
        for k, v in active_filters.items():
            entries.append((k, str(v), "field", False))
    else:
        entries.append(("(none — full dataset exported)", None, "field", False))

    entries.append((None, None, None, False))
    entries.append(("METHODOLOGY", None, "section", False))
    entries.append(("Methodology link", methodology_url, "field", True))

    entries.append((None, None, None, False))
    entries.append(("SPEC INVARIANTS", None, "section", False))
    entries.append((
        "Confidence rule",
        "COUNT(DISTINCT method_code) vs active validation_profile thresholds. "
        "Computed by the cell_triangulation_summary view — never hand-set.",
        "field", False,
    ))
    entries.append((
        "Audit trail",
        "Every number is drillable: cell → estimate → source → raw payload.",
        "field", False,
    ))
    entries.append((
        "Source invariant",
        "No cell_triangulation row exists without a non-null source_id.",
        "field", False,
    ))
    entries.append((
        "Idempotency",
        "Composite-key upserts — re-running the pipeline must not duplicate rows.",
        "field", False,
    ))

    for r_idx, (label, value, style, is_hyperlink) in enumerate(entries, start=1):
        if label is None:
            continue  # blank visual row — just advance the row counter

        a = ws.cell(row=r_idx, column=1, value=label)
        if style == "title":
            a.font = _TITLE_FONT
        elif style == "section":
            a.font = _SECTION_FONT
        else:
            a.font = _LABEL_FONT

        if value is not None:
            b = ws.cell(row=r_idx, column=2)
            if is_hyperlink:
                _set_hyperlink(b, str(value), str(value))
            else:
                b.value = value
                b.alignment = _BODY_ALIGN


# =========================================================================== #
# Shared cell-query helper
# =========================================================================== #

def _load_cells(
    session: Session,
    *,
    engagement_id: str,
    cell_id: int | None,
    subcategory_id: int | None,
    geography_id: int | None,
    year: int | None,
    confidence: str | None,
    load_triangulations: bool = False,
) -> list[Cell]:
    """Return filtered ``Cell`` rows with subcategory + geography eager-loaded.

    When *load_triangulations* is True the cell's triangulation rows are also
    eager-loaded (method + source nested), which is needed for the *cell_detail*
    and *triangulation* flavors.  For those flavors the caller pays one extra JOIN
    per loaded relationship, but avoids N+1 for the N estimates per cell.
    """
    opts = [joinedload(Cell.subcategory), joinedload(Cell.geography)]
    if load_triangulations:
        opts += [
            joinedload(Cell.triangulations).joinedload(CellTriangulation.method),
            joinedload(Cell.triangulations).joinedload(CellTriangulation.source),
        ]

    stmt = select(Cell).options(*opts).order_by(Cell.year.desc(), Cell.cell_id)

    # Engagement scope is mandatory: always filter by engagement_id so the query
    # is scoped even when no other filter is supplied.
    where = [Cell.engagement_id == engagement_id]
    if cell_id is not None:
        where.append(Cell.cell_id == cell_id)
    if subcategory_id is not None:
        where.append(Cell.subcategory_id == subcategory_id)
    if geography_id is not None:
        where.append(Cell.geography_id == geography_id)
    if year is not None:
        where.append(Cell.year == year)
    if confidence is not None:
        where.append(Cell.confidence == confidence)
    if where:
        stmt = stmt.where(*where)

    return session.execute(stmt).unique().scalars().all()


def _cell_label(cell: Cell) -> tuple[str | None, str | None, str | None]:
    """Return (subcategory_name, country, segment) display labels for *cell*."""
    return (
        cell.subcategory.name if cell.subcategory else None,
        cell.geography.country if cell.geography else None,
        cell.geography.segment if cell.geography else None,
    )


def _fmt(value: Any) -> Any:
    """Coerce Decimal and None to a spreadsheet-friendly type."""
    if isinstance(value, Decimal):
        return float(value)
    return value


# =========================================================================== #
# Flavor 1 — Cell Explorer
# =========================================================================== #

def _build_cell_explorer(
    wb: Workbook,
    session: Session,
    *,
    engagement_id: str,
    cell_id: int | None,
    subcategory_id: int | None,
    geography_id: int | None,
    year: int | None,
    confidence: str | None,
) -> None:
    """Populate a *Cells* sheet: one row per cell with TAM band + confidence chip."""
    ws = wb.create_sheet("Cells")

    headers = [
        "Cell ID", "Subcategory", "Country", "Segment", "Year",
        "TAM Revenue ($M)", "TAM Low ($M)", "TAM High ($M)", "TAM Units",
        "Confidence", "Confidence Rationale", "Status", "Last Updated",
    ]
    widths = _write_headers(ws, headers)

    cells = _load_cells(
        session,
        engagement_id=engagement_id,
        cell_id=cell_id,
        subcategory_id=subcategory_id,
        geography_id=geography_id,
        year=year,
        confidence=confidence,
    )

    CONF_COL = 10  # 1-based column index of Confidence

    for r_idx, cell in enumerate(cells, start=2):
        sub_name, country, segment = _cell_label(cell)
        row = [
            cell.cell_id,
            sub_name,
            country,
            segment,
            cell.year,
            _fmt(cell.tam_revenue_usd_m),
            _fmt(cell.tam_low_usd_m),
            _fmt(cell.tam_high_usd_m),
            cell.tam_units,
            cell.confidence,
            cell.confidence_rationale,
            cell.status,
            cell.updated_at,
        ]
        _write_row(ws, r_idx, row, widths, confidence_col=CONF_COL)

    _apply_widths(ws, widths)
    logger.debug("cell_explorer sheet: %d rows", len(cells))


# =========================================================================== #
# Flavor 2 — Cell Detail drill-down
# =========================================================================== #

def _build_cell_detail(
    wb: Workbook,
    session: Session,
    *,
    engagement_id: str,
    cell_id: int | None,
    subcategory_id: int | None,
    geography_id: int | None,
    year: int | None,
    confidence: str | None,
) -> None:
    """Two sheets: *Cell Summary* (same as Explorer) + *Estimates* (drill chain).

    The *Estimates* sheet embeds method tier, confidence cap, source publisher,
    hyperlinked source URL, and access method so the reader can follow the full
    audit trail without leaving the workbook.
    """
    cells = _load_cells(
        session,
        engagement_id=engagement_id,
        cell_id=cell_id,
        subcategory_id=subcategory_id,
        geography_id=geography_id,
        year=year,
        confidence=confidence,
        load_triangulations=True,
    )

    # --- Sheet 1: Cell Summary ---
    ws_cells = wb.create_sheet("Cell Summary")
    cell_headers = [
        "Cell ID", "Subcategory", "Country", "Segment", "Year",
        "TAM Revenue ($M)", "TAM Low ($M)", "TAM High ($M)", "TAM Units",
        "Confidence", "Confidence Rationale", "Status", "Last Updated",
    ]
    widths_c = _write_headers(ws_cells, cell_headers)
    CONF_COL_C = 10
    for r_idx, cell in enumerate(cells, start=2):
        sub_name, country, segment = _cell_label(cell)
        row = [
            cell.cell_id, sub_name, country, segment, cell.year,
            _fmt(cell.tam_revenue_usd_m), _fmt(cell.tam_low_usd_m),
            _fmt(cell.tam_high_usd_m), cell.tam_units,
            cell.confidence, cell.confidence_rationale, cell.status, cell.updated_at,
        ]
        _write_row(ws_cells, r_idx, row, widths_c, confidence_col=CONF_COL_C)
    _apply_widths(ws_cells, widths_c)

    # --- Sheet 2: Estimates (drill chain) ---
    ws_est = wb.create_sheet("Estimates")
    est_headers = [
        "Triangulation ID", "Cell ID", "Subcategory", "Country", "Segment", "Year",
        "Method Code", "Method Description", "Tier", "Confidence Cap",
        "Estimate ($M)", "Source ID", "Publisher", "Source URL",
        "Access Method", "Last Probe Status", "Notes", "Computed At",
    ]
    widths_e = _write_headers(ws_est, est_headers)
    URL_COL  = 14   # "Source URL" column (1-based)

    r_idx = 2
    for cell in cells:
        sub_name, country, segment = _cell_label(cell)
        for tri in cell.triangulations:
            method  = tri.method
            source  = tri.source
            src_url = source.url_pattern if source else None
            row = [
                tri.triangulation_id,
                cell.cell_id,
                sub_name,
                country,
                segment,
                cell.year,
                tri.method_code,
                method.description if method else None,
                method.tier if method else None,
                method.confidence_cap if method else None,
                _fmt(tri.estimate_usd_m),
                tri.source_id,
                source.publisher if source else None,
                src_url,
                source.access_method if source else None,
                source.last_probe_status if source else None,
                tri.notes,
                tri.computed_at,
            ]
            _write_row(
                ws_est, r_idx, row, widths_e,
                hyperlink_cols={URL_COL: source.publisher if source else src_url},
            )
            r_idx += 1

    _apply_widths(ws_est, widths_e)
    logger.debug("cell_detail: %d cells, %d estimates", len(cells), r_idx - 2)


# =========================================================================== #
# Flavor 3 — Player Shares
# =========================================================================== #

def _build_player_shares(
    wb: Workbook,
    session: Session,
    *,
    engagement_id: str,
    cell_id: int | None,
    subcategory_id: int | None,
    geography_id: int | None,
    year: int | None,
    confidence: str | None,
) -> None:
    """Two sheets: *Player Shares* + *Supplier Relationships*; source URLs linked.

    Cell-level filters are applied by first fetching the matching cell IDs, then
    filtering both player_shares and supplier_relationships by that set.  This
    avoids a complex multi-table join and keeps the logic readable.
    """
    # Resolve the target cell IDs from filters.
    cells = _load_cells(
        session,
        engagement_id=engagement_id,
        cell_id=cell_id,
        subcategory_id=subcategory_id,
        geography_id=geography_id,
        year=year,
        confidence=confidence,
    )
    cell_map = {c.cell_id: c for c in cells}
    cell_ids = list(cell_map.keys())

    # --- Sheet 1: Player Shares ---
    ws_ps = wb.create_sheet("Player Shares")
    ps_headers = [
        "Share ID", "Cell ID", "Subcategory", "Country", "Segment", "Year",
        "Company", "Company Type", "HQ Country",
        "Player Role", "Rank",
        "Share %", "Share Low %", "Share High %", "Revenue ($M)",
        "Source ID", "Publisher", "Source URL", "Confidence",
    ]
    widths_ps = _write_headers(ws_ps, ps_headers)
    PS_URL_COL  = 18   # "Source URL"
    PS_CONF_COL = 19   # "Confidence"

    if cell_ids:
        shares = (
            session.execute(
                select(PlayerShare)
                .options(joinedload(PlayerShare.company), joinedload(PlayerShare.source))
                .where(
                    PlayerShare.engagement_id == engagement_id,
                    PlayerShare.cell_id.in_(cell_ids),
                )
                .order_by(PlayerShare.cell_id, PlayerShare.rank)
            )
            .unique()
            .scalars()
            .all()
        )
    else:
        shares = []

    for r_idx, ps in enumerate(shares, start=2):
        cell = cell_map.get(ps.cell_id)
        sub_name, country, segment = _cell_label(cell) if cell else (None, None, None)
        co      = ps.company
        src     = ps.source
        src_url = src.url_pattern if src else None
        row = [
            ps.share_id, ps.cell_id,
            sub_name, country, segment,
            cell.year if cell else None,
            co.name if co else None,
            co.company_type if co else None,
            co.country_hq if co else None,
            ps.player_role, ps.rank,
            _fmt(ps.share_pct), _fmt(ps.share_low_pct), _fmt(ps.share_high_pct),
            _fmt(ps.revenue_usd_m),
            ps.source_id,
            src.publisher if src else None,
            src_url,
            ps.confidence,
        ]
        _write_row(
            ws_ps, r_idx, row, widths_ps,
            hyperlink_cols={PS_URL_COL: src.publisher if src else src_url},
            confidence_col=PS_CONF_COL,
        )
    _apply_widths(ws_ps, widths_ps)

    # --- Sheet 2: Supplier Relationships ---
    ws_sr = wb.create_sheet("Supplier Relationships")
    sr_headers = [
        "Relationship ID", "Buyer", "Supplier", "Cell ID",
        "Relationship Type", "Evidence Type", "Evidence Strength",
        "Source ID", "Publisher", "Source URL", "Notes",
    ]
    widths_sr = _write_headers(ws_sr, sr_headers)
    SR_URL_COL = 10  # "Source URL"

    if cell_ids:
        # SupplierRelationship has no ORM .source relationship; load buyers/suppliers
        # via joinedload and pre-fetch sources in a single IN query.
        rels = (
            session.execute(
                select(SupplierRelationship)
                .options(
                    joinedload(SupplierRelationship.buyer),
                    joinedload(SupplierRelationship.supplier),
                )
                .where(
                    SupplierRelationship.engagement_id == engagement_id,
                    SupplierRelationship.cell_id.in_(cell_ids),
                )
                .order_by(SupplierRelationship.relationship_id)
            )
            .unique()
            .scalars()
            .all()
        )
        rel_source_ids = {r.source_id for r in rels if r.source_id}
        rel_source_map: dict[str, Source] = {}
        if rel_source_ids:
            for src in (
                session.execute(
                    select(Source).where(
                        Source.engagement_id == engagement_id,
                        Source.source_id.in_(rel_source_ids),
                    )
                )
                .scalars()
                .all()
            ):
                rel_source_map[src.source_id] = src
    else:
        rels = []
        rel_source_map = {}

    for r_idx, rel in enumerate(rels, start=2):
        buyer   = rel.buyer
        supp    = rel.supplier
        src_row = rel_source_map.get(rel.source_id) if rel.source_id else None
        src_url = src_row.url_pattern if src_row else None
        row = [
            rel.relationship_id,
            buyer.name if buyer else None,
            supp.name if supp else None,
            rel.cell_id,
            rel.relationship_type,
            rel.evidence_type,
            rel.evidence_strength,
            rel.source_id,
            src_row.publisher if src_row else None,
            src_url,
            rel.notes,
        ]
        _write_row(
            ws_sr, r_idx, row, widths_sr,
            hyperlink_cols={SR_URL_COL: src_row.publisher if src_row else src_url},
        )
    _apply_widths(ws_sr, widths_sr)
    logger.debug(
        "player_shares: %d shares, %d relationships", len(shares), len(rels)
    )


# =========================================================================== #
# Flavor 4 — Triangulation
# =========================================================================== #

def _build_triangulation(
    wb: Workbook,
    session: Session,
    *,
    engagement_id: str,
    cell_id: int | None,
    subcategory_id: int | None,
    geography_id: int | None,
    year: int | None,
    confidence: str | None,
) -> None:
    """Two sheets: raw *Triangulation* estimates + *Tri Summary* confidence math.

    The *Tri Summary* sheet projects the ``cell_triangulation_summary``
    materialised view which records the ``COUNT(DISTINCT method_code)`` verdicts
    from the active validation profile.  A missing summary row means the view has
    not been refreshed or no estimates exist yet for that cell.
    """
    cells = _load_cells(
        session,
        engagement_id=engagement_id,
        cell_id=cell_id,
        subcategory_id=subcategory_id,
        geography_id=geography_id,
        year=year,
        confidence=confidence,
        load_triangulations=True,
    )
    cell_map = {c.cell_id: c for c in cells}
    cell_ids = list(cell_map.keys())

    # --- Sheet 1: Triangulation ---
    ws_tri = wb.create_sheet("Triangulation")
    tri_headers = [
        "Triangulation ID", "Cell ID", "Subcategory", "Country", "Segment", "Year",
        "Method Code", "Method Tier", "Confidence Cap",
        "Estimate ($M)", "Source ID", "Publisher", "Source URL",
        "Notes", "Computed At",
    ]
    widths_t = _write_headers(ws_tri, tri_headers)
    T_URL_COL = 13  # "Source URL"

    tri_row = 2   # separate counter so the summary loop doesn't shadow it
    for cell in cells:
        sub_name, country, segment = _cell_label(cell)
        for tri in cell.triangulations:
            method  = tri.method
            source  = tri.source
            src_url = source.url_pattern if source else None
            row = [
                tri.triangulation_id,
                cell.cell_id,
                sub_name, country, segment, cell.year,
                tri.method_code,
                method.tier if method else None,
                method.confidence_cap if method else None,
                _fmt(tri.estimate_usd_m),
                tri.source_id,
                source.publisher if source else None,
                src_url,
                tri.notes,
                tri.computed_at,
            ]
            _write_row(
                ws_tri, tri_row, row, widths_t,
                hyperlink_cols={T_URL_COL: source.publisher if source else src_url},
            )
            tri_row += 1
    _apply_widths(ws_tri, widths_t)

    # --- Sheet 2: Tri Summary (materialised view) ---
    ws_sum = wb.create_sheet("Tri Summary")
    sum_headers = [
        "Cell ID", "Subcategory", "Country", "Segment", "Year",
        "N Estimates", "N Distinct Methods", "N Independent Signals", "N Source Classes",
        "Min ($M)", "Median ($M)", "Max ($M)", "Spread Ratio",
        "Has Tier A", "Effective Signals", "Qualifies HIGH", "Qualifies MEDIUM",
    ]
    widths_s = _write_headers(ws_sum, sum_headers)

    if cell_ids:
        summaries = (
            session.execute(
                select(CellTriangulationSummary).where(
                    CellTriangulationSummary.engagement_id == engagement_id,
                    CellTriangulationSummary.cell_id.in_(cell_ids),
                )
            )
            .scalars()
            .all()
        )
    else:
        summaries = []

    for sum_row, s in enumerate(summaries, start=2):
        cell = cell_map.get(s.cell_id)
        sub_name, country, segment = _cell_label(cell) if cell else (None, None, None)
        row = [
            s.cell_id,
            sub_name, country, segment,
            cell.year if cell else None,
            s.n_estimates,
            s.n_distinct_methods,
            s.n_independent_signals,
            s.n_source_classes,
            _fmt(s.estimate_min),
            _fmt(s.estimate_median),
            _fmt(s.estimate_max),
            _fmt(s.spread_ratio),
            s.has_tier_a,
            s.effective_signals,
            s.qualifies_high,
            s.qualifies_medium,
        ]
        _write_row(ws_sum, sum_row, row, widths_s)
    _apply_widths(ws_sum, widths_s)
    logger.debug(
        "triangulation: %d estimates across %d cells, %d summary rows",
        tri_row - 2, len(cells), len(summaries),
    )


# =========================================================================== #
# Flavor 5 — Assumptions
# =========================================================================== #

def _build_assumptions(
    wb: Workbook,
    session: Session,
    *,
    engagement_id: str,
    subcategory_id: int | None,
    geography_id: int | None,
) -> None:
    """Two sheets: versioned *Assumptions* ledger + *Cell Links* bridge table.

    Filters by ``scope_subcategory_id`` / ``scope_geography_id`` when provided.
    Assumptions are never overwritten: the ``superseded_by`` chain records the
    full version history (spec invariant).  The *Cell Links* sheet provides the
    reverse-drill from each assumption to influenced cells.
    """
    # --- Load assumptions ---
    # Engagement scope is mandatory: always filter by engagement_id.
    where = [Assumption.engagement_id == engagement_id]
    if subcategory_id is not None:
        where.append(Assumption.scope_subcategory_id == subcategory_id)
    if geography_id is not None:
        where.append(Assumption.scope_geography_id == geography_id)

    stmt = select(Assumption).order_by(
        Assumption.effective_from_year.desc(), Assumption.assumption_id
    )
    if where:
        stmt = stmt.where(*where)
    assumptions = session.execute(stmt).scalars().all()

    # Pre-fetch companies referenced in scope_company_id (avoids N+1)
    company_ids = {a.scope_company_id for a in assumptions if a.scope_company_id}
    company_map: dict[int, str] = {}
    if company_ids:
        for co in (
            session.execute(
                select(Company).where(
                    Company.engagement_id == engagement_id,
                    Company.company_id.in_(company_ids),
                )
            )
            .scalars()
            .all()
        ):
            company_map[co.company_id] = co.name

    # Pre-fetch sources referenced in assumption.source_id (avoids N+1)
    source_ids = {a.source_id for a in assumptions if a.source_id}
    source_map: dict[str, Source] = {}
    if source_ids:
        for src in (
            session.execute(
                select(Source).where(
                    Source.engagement_id == engagement_id,
                    Source.source_id.in_(source_ids),
                )
            )
            .scalars()
            .all()
        ):
            source_map[src.source_id] = src

    # --- Sheet 1: Assumptions ---
    ws_ass = wb.create_sheet("Assumptions")
    ass_headers = [
        "Assumption ID", "Scope Company", "Scope Subcategory ID", "Scope Geography ID",
        "Assumption Text", "Numeric Value", "Unit", "Confidence",
        "Derivation Method", "Source ID", "Publisher", "Source URL",
        "Effective From Year", "Effective To Year", "Superseded By", "Created At",
    ]
    widths_a = _write_headers(ws_ass, ass_headers)
    A_URL_COL  = 12  # "Source URL"
    A_CONF_COL = 8   # "Confidence"

    for r_idx, a in enumerate(assumptions, start=2):
        src     = source_map.get(a.source_id) if a.source_id else None
        src_url = src.url_pattern if src else None
        row = [
            a.assumption_id,
            company_map.get(a.scope_company_id) if a.scope_company_id else None,
            a.scope_subcategory_id,
            a.scope_geography_id,
            a.assumption_text,
            _fmt(a.numeric_value),
            a.unit,
            a.confidence,
            a.derivation_method,
            a.source_id,
            src.publisher if src else None,
            src_url,
            a.effective_from_year,
            a.effective_to_year,
            a.superseded_by,
            a.created_at,
        ]
        _write_row(
            ws_ass, r_idx, row, widths_a,
            hyperlink_cols={A_URL_COL: src.publisher if src else src_url},
            confidence_col=A_CONF_COL,
        )
    _apply_widths(ws_ass, widths_a)

    # --- Sheet 2: Cell Links (reverse drill) ---
    ws_cl = wb.create_sheet("Cell Links")
    cl_headers = [
        "Cell ID", "Subcategory", "Country", "Segment", "Year",
        "TAM Revenue ($M)", "Confidence",
        "Assumption ID", "Assumption Text (preview)", "Weight",
    ]
    widths_cl = _write_headers(ws_cl, cl_headers)
    CL_CONF_COL = 7  # "Confidence"

    assumption_ids = {a.assumption_id for a in assumptions}
    if assumption_ids:
        links = (
            session.execute(
                select(CellAssumptionLink)
                .where(
                    CellAssumptionLink.engagement_id == engagement_id,
                    CellAssumptionLink.assumption_id.in_(assumption_ids),
                )
                .order_by(CellAssumptionLink.cell_id, CellAssumptionLink.assumption_id)
            )
            .scalars()
            .all()
        )
        # Pre-fetch linked cells
        link_cell_ids = list({lk.cell_id for lk in links})
        linked_cells: dict[int, Cell] = {}
        if link_cell_ids:
            for lc in (
                session.execute(
                    select(Cell)
                    .options(joinedload(Cell.subcategory), joinedload(Cell.geography))
                    .where(
                        Cell.engagement_id == engagement_id,
                        Cell.cell_id.in_(link_cell_ids),
                    )
                )
                .unique()
                .scalars()
                .all()
            ):
                linked_cells[lc.cell_id] = lc
        # Build assumption text preview lookup
        ass_text: dict[int, str] = {a.assumption_id: a.assumption_text for a in assumptions}
    else:
        links = []
        linked_cells = {}
        ass_text = {}

    for r_idx, lk in enumerate(links, start=2):
        lc = linked_cells.get(lk.cell_id)
        sub_name, country, segment = _cell_label(lc) if lc else (None, None, None)
        text_preview = (ass_text.get(lk.assumption_id) or "")[:120]
        row = [
            lk.cell_id,
            sub_name, country, segment,
            lc.year if lc else None,
            _fmt(lc.tam_revenue_usd_m) if lc else None,
            lc.confidence if lc else None,
            lk.assumption_id,
            text_preview,
            _fmt(lk.weight),
        ]
        _write_row(ws_cl, r_idx, row, widths_cl, confidence_col=CL_CONF_COL)
    _apply_widths(ws_cl, widths_cl)
    logger.debug(
        "assumptions: %d assumptions, %d cell links", len(assumptions), len(links)
    )


# =========================================================================== #
# Public dispatcher
# =========================================================================== #

def build_workbook(
    flavor: ExportFlavor,
    session: Session,
    *,
    engagement_id: str,
    cell_id: int | None = None,
    subcategory_id: int | None = None,
    geography_id: int | None = None,
    year: int | None = None,
    confidence: str | None = None,
) -> Workbook:
    """Build and return the openpyxl ``Workbook`` for the requested export *flavor*.

    The workbook always starts with a ``_README`` sheet at position 0 (added after
    the data sheets by moving it forward) so it is the first tab the user sees.

    Parameters
    ----------
    flavor:
        ``cell_explorer`` | ``cell_detail`` | ``player_shares`` |
        ``triangulation`` | ``assumptions``
    session:
        Active SQLAlchemy session.  Does not manage transactions.
    cell_id:
        When provided, scopes ``cell_detail`` to exactly one cell.  Ignored by
        ``assumptions``.
    subcategory_id / geography_id / year / confidence:
        Cell-level filters; applied by all flavors that operate on cells.  The
        ``assumptions`` flavor uses ``subcategory_id`` and ``geography_id`` as
        scope filters on ``scope_subcategory_id`` / ``scope_geography_id``.

    Returns
    -------
    openpyxl.Workbook
        Caller should pass it to :func:`workbook_to_bytes` and stream the result.

    Raises
    ------
    ValueError
        When *flavor* is not one of the five accepted values.
    """
    wb = Workbook()
    # Remove the default empty sheet openpyxl always creates.
    if wb.active:
        wb.remove(wb.active)

    ts = datetime.now(timezone.utc)
    methodology_url = f"{settings.NEXT_PUBLIC_APP_URL.rstrip('/')}/methodology"

    # Build filter description for the _README sheet.
    filter_desc: dict[str, Any] = {
        "cell_id": cell_id,
        "subcategory_id": subcategory_id,
        "geography_id": geography_id,
        "year": year,
        "confidence": confidence,
    }

    if flavor == "cell_explorer":
        _build_cell_explorer(
            wb, session,
            engagement_id=engagement_id,
            cell_id=cell_id,
            subcategory_id=subcategory_id,
            geography_id=geography_id,
            year=year,
            confidence=confidence,
        )

    elif flavor == "cell_detail":
        _build_cell_detail(
            wb, session,
            engagement_id=engagement_id,
            cell_id=cell_id,
            subcategory_id=subcategory_id,
            geography_id=geography_id,
            year=year,
            confidence=confidence,
        )

    elif flavor == "player_shares":
        _build_player_shares(
            wb, session,
            engagement_id=engagement_id,
            cell_id=cell_id,
            subcategory_id=subcategory_id,
            geography_id=geography_id,
            year=year,
            confidence=confidence,
        )

    elif flavor == "triangulation":
        _build_triangulation(
            wb, session,
            engagement_id=engagement_id,
            cell_id=cell_id,
            subcategory_id=subcategory_id,
            geography_id=geography_id,
            year=year,
            confidence=confidence,
        )

    elif flavor == "assumptions":
        _build_assumptions(
            wb, session,
            engagement_id=engagement_id,
            subcategory_id=subcategory_id,
            geography_id=geography_id,
        )
        # assumptions flavor does not use cell_id / year / confidence
        filter_desc = {
            "scope_subcategory_id": subcategory_id,
            "scope_geography_id": geography_id,
        }

    else:
        raise ValueError(
            f"Unknown export flavor {flavor!r}. "
            "Must be one of: cell_explorer, cell_detail, player_shares, "
            "triangulation, assumptions."
        )

    # Add _README as the first sheet (position 0).
    _add_readme_sheet(wb, flavor, filter_desc, ts, methodology_url)

    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    """Serialize *wb* to raw XLSX bytes suitable for a ``StreamingResponse``."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


__all__ = [
    "ExportFlavor",
    "build_workbook",
    "workbook_to_bytes",
]
